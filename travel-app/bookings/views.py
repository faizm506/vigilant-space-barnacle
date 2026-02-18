import csv
from django.http import HttpResponse
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from .models import Booking # We will create this model next
from django.db.models import Sum, Q


@login_required
def dashboard(request):
    query = request.GET.get('search')
    
    if query:
        bookings = Booking.objects.filter(
            Q(customer_name__icontains=query) | Q(booking_id__icontains=query)
        ).order_by('-id')
    else:
        bookings = Booking.objects.all().order_by('-id')

    # Calculate Totals
    total_travelers = bookings.aggregate(Sum('total_members'))['total_members__sum'] or 0
    unpaid_count = bookings.exclude(payment_status='Paid').count()

    context = {
        'bookings': bookings,
        'total_travelers': total_travelers,
        'unpaid_count': unpaid_count,
    }

    # THE FIX: If the request is from HTMX, render ONLY the rows
    if request.headers.get('HX-Request'):
        return render(request, 'bookings/partials/booking_rows.html', context)
    
    # Otherwise, render the full dashboard page
    return render(request, 'bookings/dashboard.html', context)

from django.contrib import messages
from django.db import transaction

@login_required
def new_booking(request):
    if request.method == 'POST':
        try:
            # 1. Capture and Validate raw inputs
            # Using try/except blocks ensures the app doesn't crash on bad input
            try:
                price_per_person = float(request.POST.get('tour_price', 0) or 0)
                members_count = int(request.POST.get('total_members', 1) or 1)
            except ValueError:
                messages.error(request, "Invalid numeric input for price or members.")
                return render(request, 'bookings/booking_form.html')

            # 2. Automatically calculate the Total Price
            calculated_total = price_per_person * members_count

            # 3. Create the record inside an atomic transaction
            # This ensures that if the file upload fails, the database entry isn't half-saved
            with transaction.atomic():
                booking = Booking.objects.create(
                    customer_name=request.POST.get('customer_name'),
                    address=request.POST.get('address'),
                    total_members=members_count,
                    passport_photo=request.FILES.get('passport_photo'),
                    tour_price=calculated_total,
                    payment_status=request.POST.get('payment_status'),
                    # Enhanced JSON Data with Timestamp and Manager Info
                    additional_info={
                        "Price Per Person": f"₹{price_per_person}",
                        "Meal Preference": request.POST.get('meal_pref'),
                        "Hotel Grade": request.POST.get('hotel_stars'),
                        "Created By": request.user.username,
                        "Entry Method": "Standard Portal"
                    }
                )

            # 4. Success Feedback
            messages.success(request, f"Voucher {booking.booking_id} generated for {booking.customer_name}.")
            return redirect('dashboard')

        except Exception as e:
            # Catch-all for unexpected errors (e.g., database connection, file system issues)
            messages.error(request, f"An unexpected error occurred: {str(e)}")
            
    return render(request, 'bookings/booking_form.html')


from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from django.http import HttpResponse

@login_required
def export_bookings_csv(request):
    # Create workbook and active sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Razak and Sons Bookings"

    # 1. Define Premium Headers
    headers = ['Booking UID', 'Customer Name', 'Total Members', 'Package Price', 'Payment Status', 'Date Booked']
    ws.append(headers)

    # Styling the header row (Bold, White Text, Navy Background)
    header_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # 2. Add Data Rows
    bookings = Booking.objects.all().order_by('-created_at')
    for b in bookings:
        ws.append([
            b.booking_id,
            b.customer_name,
            b.total_members,
            float(b.tour_price), # Ensure it's a number for Excel math
            b.payment_status,
            b.created_at.replace(tzinfo=None) # Excel doesn't like timezones
        ])

    # 3. Auto-Adjust Column Widths
    # This loop calculates the length of the longest value in each column
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # 4. Return the response as an Excel file
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="Razak_Sons_Bookings.xlsx"'
    
    wb.save(response)
    return response

from django.shortcuts import render, get_object_or_404

@login_required
def booking_detail(request, booking_id):
    # This fetches the specific booking or shows a 404 error if not found
    booking = get_object_or_404(Booking, booking_id=booking_id)
    return render(request, 'bookings/booking_detail.html', {'booking': booking})

from django.shortcuts import get_object_or_404, redirect
from django.contrib import messages

@login_required
def delete_booking(request, booking_id):
    booking = get_object_or_404(Booking, booking_id=booking_id)
    if request.method == 'POST':
        customer_name = booking.customer_name
        booking.delete()
        # Adding a success message for the manager
        messages.success(request, f"Booking for {customer_name} has been successfully removed.")
    return redirect('dashboard')


from django.db.models import Sum

@login_required
def batch_export_view(request):
    start_date = request.GET.get('from_date')
    end_date = request.GET.get('to_date')
    
    bookings = None
    total_revenue = 0
    total_pax = 0
    
    if start_date and end_date:
        bookings = Booking.objects.filter(
            created_at__date__range=[start_date, end_date]
        ).order_by('created_at')
        
        # Calculate Totals for the Insight Strip
        # We use .aggregate to perform math directly in the database for speed
        stats = bookings.aggregate(
            rev_sum=Sum('tour_price'),
            pax_sum=Sum('total_members')
        )
        total_revenue = stats['rev_sum'] or 0
        total_pax = stats['pax_sum'] or 0

    context = {
        'bookings': bookings,
        'from_date': start_date,
        'to_date': end_date,
        'total_revenue': f"{total_revenue:,}", # Formats with commas like 1,00,000
        'total_pax': total_pax,
    }
    return render(request, 'bookings/batch_export.html', context)


import io
import zipfile
from django.template.loader import render_to_string
from django.http import HttpResponse
from weasyprint import HTML
from .models import Booking

@login_required
def export_batch_zip(request):
    start_date = request.GET.get('from_date')
    end_date = request.GET.get('to_date')
    
    if not start_date or not end_date:
        return HttpResponse("Please select a valid date range.", status=400)

    bookings = Booking.objects.filter(created_at__date__range=[start_date, end_date])

    if not bookings.exists():
        return HttpResponse("No records found for this period.", status=404)

    zip_buffer = io.BytesIO()

    with zipfile.ZipFile(zip_buffer, 'w') as zip_file:
        for booking in bookings:
            # We pass the full request to render_to_string to handle absolute paths
            html_string = render_to_string('bookings/booking_detail.html', {
                'booking': booking,
                'is_pdf': True # Optional flag if you want to hide buttons in the PDF
            }, request=request)
            
            pdf_file = io.BytesIO()
            # base_url ensures that your CSS and static files are found correctly
            HTML(string=html_string, base_url=request.build_absolute_uri('/')).write_pdf(pdf_file)
            
            # Sanitize filename: remove spaces/special characters from customer name
            safe_name = "".join([c for c in booking.customer_name if c.isalnum() or c==' ']).rstrip().replace(' ', '_')
            file_name = f"{booking.booking_id}_{safe_name}.pdf"
            
            zip_file.writestr(file_name, pdf_file.getvalue())

    zip_buffer.seek(0)
    response = HttpResponse(zip_buffer.read(), content_type='application/zip')
    
    # Branded filename for the ZIP itself
    download_name = f"RazakSons_Backup_{start_date}_to_{end_date}.zip"
    response['Content-Disposition'] = f'attachment; filename="{download_name}"'
    
    return response